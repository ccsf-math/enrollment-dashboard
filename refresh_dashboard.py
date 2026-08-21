#!/usr/bin/env python3
"""Refresh the enrollment dashboard's baked-in data.

Run this after you update "Fall 2026 Enrollment.xlsx" or drop a new CSV
snapshot into data/. It rewrites index.html in place with:

  * the daily enrollment history from the workbook, and
  * every CSV snapshot in data/ (plus a matching data/manifest.json).

You do NOT need this just to see fresh trend lines: dragging the workbook
onto the open dashboard updates the daily history straight in the browser.
The script matters when you want the change baked into the file itself, or
when you have added new CSV snapshots (those carry capacity, instructor and
room, which the workbook does not).

Usage, from inside the "enrollment-dashboard" repository:

    python3 refresh_dashboard.py

Requires openpyxl:  pip3 install openpyxl
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip3 install openpyxl")

HERE = Path(__file__).parent
DASHBOARD = HERE / "index.html"
DATA_DIR = HERE / "data"

DATA_START = "/* DAILY-DATA-START */"
DATA_END = "/* DAILY-DATA-END */"
FALL_START = "/* FALL-HISTORY-START */"
FALL_END = "/* FALL-HISTORY-END */"

MAX_COMPARE_DAYS = 151   # registration open through roughly a month into the term

# Courses that have been renumbered, mapped onto the number in use today so a
# course keeps one continuous line across terms. Statistics moved out of MATH
# in Fall 2025: MATH 80 -> STAT C1000, MATH 80S -> STAT 80S.
COURSE_ALIASES = {
    ("STAT", "80"): ("STAT", "C1000"),
    ("MATH", "80"): ("STAT", "C1000"),
    ("MATH", "80S"): ("STAT", "80S"),
}
LABEL_RE = re.compile(r"^([A-Za-z]+)\s+(\S+)-(\S+)$")

# ---- FTEF, exactly as the dashboard budget card computes it ----

# Credit hours per section, matching COURSE_UNITS in the dashboard.
UNITS = {
    'MATH 65': 3, 'MATH 70': 3, 'MATH 75': 3, 'MATH 75S': 2, 'MATH 90': 5,
    'MATH 90S': 3, 'MATH 95': 3, 'MATH 98': 7, 'MATH 100A': 3, 'MATH 100B': 3,
    'MATH 108': 5, 'MATH 110A': 5, 'MATH 110B': 5, 'MATH 110C': 5, 'MATH 110S': 0,
    'MATH 115': 3, 'MATH 120': 3, 'MATH 125': 3, 'MATH 130': 5, 'MATH 9000': 0.857,
    'STAT C1000': 5, 'STAT 80S': 2, 'STAT 9000': 0.857,
}

# primary -> support. The support half folds into the primary.
PAIRS = {'MATH 90': 'MATH 90S', 'MATH 75': 'MATH 75S',
         'MATH 110A': 'MATH 110S', 'STAT C1000': 'STAT 80S'}
SUPPORT = {v: k for k, v in PAIRS.items()}

LCM_THRESHOLD = 60
LCM_MULTIPLIER = 1.5
COST_DIVISOR = 15


def ftef_for_day(enrolled_by_section):
    """enrolled_by_section: {(course_code, section): enrolled}. Returns a dict."""
    live = {k: v for k, v in enrolled_by_section.items() if v is not None and v > 0}

    weighted = 0.0
    base = 0.0
    sections = 0
    paired = 0
    lcm = 0
    enrolled = 0

    for (code, sec), enr in live.items():
        # Skip a support half whose primary is running; it folds into the primary.
        if code in SUPPORT and (SUPPORT[code], sec) in live:
            continue

        is_paired = code in PAIRS and (PAIRS[code], sec) in live
        units = UNITS.get(code)
        is_lcm = enr >= LCM_THRESHOLD

        sections += 1
        paired += is_paired
        lcm += is_lcm
        enrolled += enr

        if units is None:
            continue
        base += units
        weighted += units * (LCM_MULTIPLIER if is_lcm else 1)

    return {
        'ftef': round(weighted / COST_DIVISOR, 4),
        'sections': sections,
        'paired': paired,
        'lcm': lcm,
        'baseUnits': round(base, 3),
        'weightedUnits': round(weighted, 3),
        'enrolled': enrolled,
    }



def split_label(label):
    """"MATH  65-001" -> ("MATH", "65", "001"), with renumbering applied."""
    m = LABEL_RE.match(re.sub(r"\s+", " ", str(label)).strip())
    if not m:
        return None
    subject, course = COURSE_ALIASES.get(
        (m.group(1).upper(), m.group(2)), (m.group(1).upper(), m.group(2)))
    return subject, course, m.group(3)


def normalize_key(label):
    """Section key as the dashboard writes it: Subject|Course|Section."""
    parts = split_label(label)
    return "|".join(parts) if parts else None


def term_from_filename(name):
    # Matches both "Fall 2026 Enrollment.xlsx" and "enrollment_Fall-2026_....csv"
    m = re.search(r"(Fall|Spring|Summer|Winter)[\s_-]*(\d{4})", name, re.I)
    return f"{m.group(1).capitalize()} {m.group(2)}" if m else None


def read_workbook(path):
    """Grid whose first row is dates and first column is section labels."""
    ws = openpyxl.load_workbook(path, data_only=True, read_only=True).worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        raise SystemExit(f"{path.name} is empty")

    dates, cols = [], []
    for idx, cell in enumerate(rows[0][1:], start=1):
        if isinstance(cell, datetime):
            dates.append(cell.strftime("%Y-%m-%d"))
            cols.append(idx)
        elif isinstance(cell, (int, float)):
            # Excel serial date, 1900 system
            iso = datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(cell))
            dates.append(iso.strftime("%Y-%m-%d"))
            cols.append(idx)
    if len(dates) < 2:
        raise SystemExit(f"{path.name}: no date columns found in the first row")

    series = {}
    for row in rows[1:]:
        key = normalize_key(row[0]) if row[0] else None
        if not key:
            continue    # header repeats and the "Total by COLUMNS" row land here
        values = [int(row[c]) if c < len(row) and isinstance(row[c], (int, float)) else None
                  for c in cols]
        if any(v is not None for v in values):
            series[key] = values
    if not series:
        raise SystemExit(f'{path.name}: no section rows recognized (expected labels like "MATH  65-001")')

    return {"dates": dates, "series": series}


def course_history(path):
    """Course-level daily totals for the Fall-to-Fall comparison.

    Readings are indexed by calendar days since registration opened rather than
    by column position, because the workbooks skip the odd day. Missing days
    stay null and the dashboard carries the last reading forward.
    """
    ws = openpyxl.load_workbook(path, data_only=True, read_only=True).worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return None

    dated = []
    for idx, cell in enumerate(rows[0][1:], start=1):
        if isinstance(cell, datetime):
            dated.append((idx, cell.date()))
        elif isinstance(cell, (int, float)):
            dated.append((idx, datetime.fromordinal(
                datetime(1899, 12, 30).toordinal() + int(cell)).date()))
    if len(dated) < 2:
        return None

    start = dated[0][1]
    offsets = [(c, (d - start).days) for c, d in dated if 0 <= (d - start).days < MAX_COMPARE_DAYS]
    if not offsets:
        return None
    span = max(off for _, off in offsets) + 1

    # section (course, sec) -> reading per day offset
    sections = {}
    for row in rows[1:]:
        parts = split_label(row[0]) if row[0] else None
        if not parts:
            continue
        by_off = sections.setdefault((f"{parts[0]} {parts[1]}", parts[2]), {})
        for c, off in offsets:
            v = row[c] if c < len(row) else None
            if isinstance(v, (int, float)):
                by_off[off] = int(v)
    if not sections:
        return None

    courses, ftef, sec_count, enrolled = {}, [None] * span, [None] * span, [None] * span
    for _, off in offsets:
        day = {k: v.get(off) for k, v in sections.items()}
        if all(v is None for v in day.values()):
            continue
        for (code, _sec), v in day.items():
            if v is None:
                continue
            series = courses.setdefault(code, [None] * span)
            series[off] = (series[off] or 0) + v
        m = ftef_for_day(day)
        ftef[off] = m["ftef"]
        sec_count[off] = m["sections"]
        enrolled[off] = m["enrolled"]

    return {
        "start": start.strftime("%Y-%m-%d"),
        "courses": courses,
        "ftef": ftef,
        "sections": sec_count,
        "enrolled": enrolled,
    }


def main():
    if not DASHBOARD.exists():
        sys.exit(f"{DASHBOARD.name} not found next to this script")
    html = DASHBOARD.read_text()
    if DATA_START not in html:
        sys.exit(f"{DASHBOARD.name} has no daily-history block; this script expects the updated dashboard")

    # Terms the dashboard shows section cards for. Only these need section-level
    # daily detail; prior years are carried as course totals, which is all the
    # Fall-to-Fall section uses and a fraction of the size.
    live_terms = {t for t in (term_from_filename(p.name) for p in DATA_DIR.glob("*.csv")) if t}

    # ---- daily history from every workbook in data/
    payload, fall_hist = {}, {}
    for xlsx in sorted(DATA_DIR.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue    # Excel lock file
        term = term_from_filename(xlsx.name)
        if not term:
            print(f"  skipped {xlsx.name}: cannot tell which term it covers")
            continue

        if term in live_terms:
            parsed = read_workbook(xlsx)
            payload[term] = parsed
            print(f"  {term}: {len(parsed['series'])} sections, {len(parsed['dates'])} days "
                  f"({parsed['dates'][0]} to {parsed['dates'][-1]})")

        if term.startswith("Fall "):
            hist = course_history(xlsx)
            if hist:
                fall_hist[term] = hist
                if term not in live_terms:
                    n = len(next(iter(hist["courses"].values())))
                    print(f"  {term}: {len(hist['courses'])} courses over {n} days "
                          f"from {hist['start']} (comparison only)")

    # ---- course-level history behind the Fall-to-Fall section
    if fall_hist and FALL_START in html:
        fall_blob = json.dumps(fall_hist, separators=(",", ":"))
        html = re.sub(
            re.escape(FALL_START) + ".*?" + re.escape(FALL_END),
            lambda _: f"{FALL_START}\nwindow.EMBEDDED_FALL_HISTORY = {fall_blob};\n{FALL_END}",
            html, flags=re.S,
        )
        print(f"  Fall-to-Fall comparison: {', '.join(sorted(fall_hist))}")

    if payload:
        blob = json.dumps(payload, separators=(",", ":"))
        html = re.sub(
            re.escape(DATA_START) + ".*?" + re.escape(DATA_END),
            lambda _: f"{DATA_START}\nwindow.EMBEDDED_DAILY = {blob};\n{DATA_END}",
            html, flags=re.S,
        )
    else:
        print("  no workbooks found in data/; daily history left as it was")

    # ---- CSV snapshots
    csvs = sorted(DATA_DIR.glob("*.csv"))
    if csvs:
        embedded = json.dumps({p.name: p.read_text() for p in csvs}, separators=(",", ":"))
        pattern = re.compile(r"<script>var EMBEDDED_CSVS = \{.*?\};</script>", re.S)
        if not pattern.search(html):
            sys.exit("could not find the embedded CSV block in the dashboard")
        html = pattern.sub(lambda _: f"<script>var EMBEDDED_CSVS = {embedded};</script>", html, count=1)
        (DATA_DIR / "manifest.json").write_text(
            json.dumps({"files": [p.name for p in csvs]}, indent=2) + "\n")
        print(f"  {len(csvs)} CSV snapshots embedded (latest {csvs[-1].name})")
    else:
        print("  no CSVs found in data/; snapshots left as they were")

    DASHBOARD.write_text(html)
    print(f"\nUpdated {DASHBOARD.name} ({DASHBOARD.stat().st_size / 1e6:.1f} MB).")
    print("If the dashboard is open, reload the page. Your saved budget settings are untouched.")


if __name__ == "__main__":
    main()
